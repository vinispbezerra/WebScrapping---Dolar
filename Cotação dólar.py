import requests
from bs4 import BeautifulSoup
from datetime import datetime
import openpyxl


def obter_cotacao_dolar():
    url = 'https://dolarhoje.com'
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3'}

    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()  # Lança uma exceção para erros HTTP
        soup = BeautifulSoup(response.text, 'html.parser')

        # Tente encontrar o elemento de cotação de diferentes maneiras
        cotacao_dolar = soup.find('input', {'id': 'nacional'}).get('value')
        if not cotacao_dolar:
            cotacao_dolar = soup.find('input', {'id': 'comercial'}).get('value')
        if not cotacao_dolar:
            cotacao_dolar = soup.find('input', {'id': 'turismo'}).get('value')

        return cotacao_dolar
    except requests.exceptions.RequestException as e:
        print(f'Erro ao obter dados: {e}')
        return None

def adicionar_cotacao_a_planilha(cotacao_dolar):
    nome_arquivo = 'cotacao_dolar.xlsx'

    try:
        # Tenta abrir uma planilha existente
        wb = openpyxl.load_workbook(nome_arquivo)
        ws = wb.active
    except FileNotFoundError:
        # Se a planilha não existir, cria uma nova
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'Data'
        ws['B1'] = 'Hora'
        ws['C1'] = 'Cotação do Dólar'
        ws['D1'] = 'Média'

    # Verifica se a cotação_dolar é um número
    try:
        cotacao_float = float(cotacao_dolar.replace(',', '.'))
    except ValueError:
        print(f'Cotação do dólar não é um valor numérico. Não adicionando à planilha.')
        return

    # Adiciona uma nova linha com a data, hora atual e a cotação do dólar
    proxima_linha = ws.max_row + 1
    ws.cell(row=proxima_linha, column=1, value=datetime.now().date())
    ws.cell(row=proxima_linha, column=2, value=datetime.now().strftime('%H:%M:%S'))
    ws.cell(row=proxima_linha, column=3, value=cotacao_float)

    # Calcula a média considerando apenas valores não vazios ou None
    valores_cotacao = [float(cell.value) for cell in ws['C'][1:proxima_linha] if cell.value is not None and cell.value != '']
    media = sum(valores_cotacao) / len(valores_cotacao) if valores_cotacao else 0

    # Adiciona a média à célula correspondente
    ws.cell(row=proxima_linha, column=4, value=media)

    # Salva a planilha
    wb.save(nome_arquivo)
    print('Cotação adicionada à planilha com sucesso.')


    # Adiciona a média à célula correspondente
    ws.cell(row=proxima_linha, column=4, value=media)
if __name__ == '__main__':
    cotacao = obter_cotacao_dolar()

    if cotacao:
        print(f'Cotação do dólar hoje: {cotacao}')
        adicionar_cotacao_a_planilha(cotacao)
    else:
        print('Erro ao obter cotação do dólar.')
